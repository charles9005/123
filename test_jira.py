from jira import JIRA
from configparser import ConfigParser
#取配置值
conf = ConfigParser()
conf.read("jira_item.ini",encoding="utf-8")
userEmail = conf.get("jirainfo","userEmail")
api_token = conf.get("jirainfo","api_token")
reporter = conf.get("jql","reporter")
project_key = conf.get("project","project_key")
jira_server =conf .get("jirainfo","jira_server")
if reporter == "()":
    jql = f'project = {project_key} AND status not in ("CLOSED") AND issuetype = Bug ORDER BY created DESC'
else:
    jql =f'project = {project_key} AND reporter in {reporter} AND status not in ("CLOSED") AND issuetype = Bug ORDER BY created DESC'

print(jql)

#Should keep api_token as secret
jira = JIRA(server=jira_server,basic_auth=(userEmail, api_token))
#默认50行，手动修改用来调试
issue_lists = jira.search_issues(jql,maxResults=1000)
#获取bug的总行数
issue_num = len(issue_lists)
# print(type(issue_lists))

#project = jira.project('某个项目的Key')

project = jira.project(project_key)
project_name =project.name
print("当前正在打印",project_name,"项目的bug汇总信息，请等待片刻")

#Excel操作
import os
file_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(file_dir,"Test status of projects.xlsx")
# 1、加载excel数据文件
from openpyxl import load_workbook
wb = load_workbook(file_path)
# 2、根据表单名称选择表单：wb['表单名称']
sh = wb["bugs"]

# 1、拿到字典的key值：
# print(list(sh.rows)[0])  # (<Cell 'login'.A1>, <Cell 'login'.B1>, <Cell 'login'.C1>)
titles = []
for item in list(sh.rows)[0]: # 遍历第1行当中每一列
    titles.append(item.value)
# print(titles)


# issue list
total_list = []
for issue in issue_lists:
    # bug url link
    bug_url = f'https://jabra1.atlassian.net/browse/{issue.key}'
    bug = jira.issue(issue.key)
    # 打印输出信息
    bug_key,bug_summary,bug_priority,bug_url =issue.key,str(bug.fields.summary),str(bug.fields.priority),str(bug_url)


    bug_list =[bug_key,bug_summary,bug_priority,bug_url]
    # print(bug_key,bug_summary,bug_priority,bug_url,sep="  ")
    total_list.append(bug_list)
#print(total_list)
# 写入第二行第1,2,3,4
# print(total_list[0][0])
# print(total_list[0][1])
# print(total_list[0][2])
# print(total_list[0][3])

for i in range(2,issue_num+2):
    for j in range(1,5):
        sh.cell(i,j).value = total_list[i-2][j-1]
        sh.cell(i, 4).value ='=HYPERLINK("{}","{}")'.format(total_list[i-2][3],"Click to broswer in JAMA")
#bug_url设置超链接
# 保存数据
import datetime
nowday = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
wb.save(os.path.join(file_dir,"Test status of {0} project_{1}.xlsx".format(project_name,nowday)))
print("文件写入成功，请到根目录下查看对应日期的文件")







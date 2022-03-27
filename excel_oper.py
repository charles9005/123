import os
file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),"Test status of projects.xlsx")
print(file_path)

# 1、加载excel数据文件
from openpyxl import load_workbook
wb = load_workbook(file_path)

# 2、根据表单名称选择表单：wb['表单名称']
sh = wb["bugs"]
#读取excel表头数据
# print(sh.cell(1, 1).value)

#向excel中写入数据
import datetime
nowday = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
print(nowday)



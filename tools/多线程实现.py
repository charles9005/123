from jira import JIRA
from configparser import ConfigParser
import queue
import threading
import re
import time
from multiprocessing.pool import ThreadPool
from openpyxl import Workbook


def set_excel_format():
    # create a new workbook
    wb = Workbook()
    sh_old = wb["Sheet"]
    wb.remove(sh_old)
    sh = wb.create_sheet('Issues')
    sh.cell(1, 1, 'Issue_Key')
    sh.cell(1, 2, 'Summary')
    sh.cell(1, 3, 'Assignee')
    sh.cell(1, 4, 'Reporter')
    sh.cell(1, 5, 'Priority')
    sh.cell(1, 6, 'Status')
    # 设置summary行宽
    sh.column_dimensions['B'].width = 80


def downloader(q: queue.Queue):
    while not q.empty():
        print(q.get())
        q.task_done()


class JiraNew():
    # 初始化获取配置信息以及连接jira
    def __init__(self, conf_filepath):
        conf = ConfigParser()
        self.conf_filepath = conf_filepath
        conf.read(self.conf_filepath, encoding="utf-8")
        self.userEmail = conf.get("jirainfo", "userEmail")
        self.api_token = conf.get("jirainfo", "api_token")
        self.reporter = conf.get("jql", "reporter")
        self.project_key = conf.get("project", "project_key")
        self.jira_server = conf.get("jirainfo", "jira_server")

    def get_jql(self):
        if self.reporter == "()":
            jql = f'project = {self.project_key} AND status not in ("CLOSED") AND issuetype = Bug ORDER BY created DESC'
        else:
            jql = f'project = {self.project_key} AND reporter in {self.reporter} AND status not in ("CLOSED") AND issuetype = Bug ORDER BY created DESC'
        return jql

    def connect_jira(self):
        jira = JIRA(server=self.jira_server, basic_auth=(self.userEmail, self.api_token))
        return jira

    def get_issues_list(self):
        return self.connect_jira().search_issues(self.get_jql(), maxResults=1000)

    def get_issues_num(self):
        return len(self.get_issues_list())

    def issues_infos(self):
        pass

    def write_issues_to_excel(self):
        pass


if __name__ == '__main__':
    start_time = time.time()
    conf_filepath = 'jira_item.ini'
    jn = JiraNew(conf_filepath)
    q = queue.Queue()
    num = jn.get_issues_num()
    # print(num)
    for i in range(num):
        q.put(f'{jn.jira_server}/browse/{jn.get_issues_list()[i]}')
    pool = ThreadPool(3)
    pool.apply_async(downloader, args=(q,))
    q.join()
    end_time = time.time()
    print(f'finished,使用时间{end_time - start_time}秒')

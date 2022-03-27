'''
Author:Charles Lin
Version:0.00
Tool to download issue list direct from jira
'''

import time
from jira import JIRA
from configparser import ConfigParser
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import threading


class JiraTest:
    def __init__(self, conf_filepath):
        conf = ConfigParser()
        self.conf_filepath = conf_filepath
        conf.read(self.conf_filepath, encoding="utf-8")
        self.userEmail = conf.get("jirainfo", "userEmail")
        self.api_token = conf.get("jirainfo", "api_token")
        self.reporter = conf.get("jql", "reporter")
        self.project_key = conf.get("project", "project_key").upper()
        self.jira_server = conf.get("jirainfo", "jira_server")
        if self.reporter == "()":
            self.jql = f'project = {self.project_key} AND status not in ("CLOSED") AND issuetype = Bug ORDER BY created DESC'
        else:
            self.jql = f'project = {self.project_key} AND reporter in {self.reporter} AND status not in ("CLOSED") AND issuetype = Bug ORDER BY created DESC'

    def connect_jira(self):
        jira = JIRA(server=self.jira_server, basic_auth=(self.userEmail, self.api_token))
        return jira

    def get_issues(self):
        # maxResults默认50行，手动修改用来调试
        issue_lists = self.connect_jira().search_issues(self.jql, maxResults=500)
        return issue_lists

    def get_issue_len(self):
        return len(self.get_issues())

    def get_project_info(self):
        project_name = self.connect_jira().project(self.project_key).name
        return project_name

    def print_project_info(self):
        project_name = self.get_project_info()
        print("正在下载", project_name, "项目的bug汇总信息，请等待...")

    def set_excel_format(self):
        pass

    def get_issues_list(self):
        pass

    def write_issue_to_excel(self):
        pass

    def create_output_excel(self):
        # create a new workbook
        wb = Workbook()
        sh_old = wb["Sheet"]
        wb.remove(sh_old)
        sh = wb.create_sheet('Issues')
        f1 = Font(name="微软雅黑", size=11, bold=True, italic=False, color="180018")
        font_red = Font(color="FB0903")
        sh.cell(1, 1, 'Issue_Key').font = f1
        sh.cell(1, 2, 'Summary').font = f1
        sh.cell(1, 3, 'Assignee').font = f1
        sh.cell(1, 4, 'Reporter').font = f1
        sh.cell(1, 5, 'Priority').font = f1
        sh.cell(1, 6, 'Status').font = f1
        # 设置summary行宽
        sh.column_dimensions['A'].width = 10
        sh.column_dimensions['B'].width = 80
        sh.column_dimensions['C'].width = 13
        sh.column_dimensions['D'].width = 13

        # issue list
        total_list = []
        for issue in self.get_issues():
            # bug url link
            bug_url = f'https://jabra1.atlassian.net/browse/{issue.key}'
            bug = self.connect_jira().issue(issue.key)
            # 打印输出信息
            bug_key, bug_summary, bug_assignee, bug_reporter, bug_priority, bug_status = issue.key, str(
                bug.fields.summary), str(bug.fields.assignee), str(bug.fields.reporter), str(
                str(bug.fields.priority)), str(bug.fields.status)
            bug_url = str(bug_url)
            bug_list = [bug_key, bug_summary, bug_assignee, bug_reporter, bug_priority, bug_status, bug_url]
            total_list.append(bug_list)

        # 写入jira数据到表格中
        for i in range(2, self.get_issue_len() + 2):
            for j in range(1, 7):
                sh.cell(i, j).value = total_list[i - 2][j - 1]
            if sh.cell(i, 5).value in ['Blocker', 'Critical', 'Major']:
                sh.cell(i, 5).font = font_red

            # 设置超链接
            sh.cell(i, 1).hyperlink = total_list[i - 2][6]
        timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        wb_name = "Test_status_of_{0}_project_{1}.xlsx".format(self.get_project_info(), timestamp).replace(" ", "_")
        wb.save(wb_name)
        wb.close()
        print("即将打开文件，请稍后...")
        self.open_output_file(wb_name)

    def open_output_file(self, file_name):
        try:
            cmd = "start /MAX excel " + r'' + file_name
            os.system(cmd)
        except:
            print("Fail to start excel application")


if __name__ == '__main__':
    start_time = time.time()
    # conf_filepath = sys.argv[1]
    conf_filepath = 'jira_item.ini'
    jt = JiraTest(conf_filepath)

    # threading
    t1 = threading.Thread(target=)
    t2 = threading.Thread(target=)
    t3 = threading.Thread(target=)

    t1.start()
    t2.start()
    t3.start()

    t1.join()
    t2.join()
    t3.join()

    jt.print_project_info()
    jt.create_output_excel()
    end_time = time.time()
    print(f'本次运行耗时{end_time - start_time}秒')
